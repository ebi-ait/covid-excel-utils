from contextlib import closing

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from .validate import ValidatingExcel


class ExcelMarkup(ValidatingExcel):
    def __init__(self, excel_path: str, sheet_index=0):
        self.__clear_markup(excel_path, sheet_index)
        super().__init__(excel_path, sheet_index)
        self.__path = excel_path
        self.__sheet_index = sheet_index
        self.__book = load_workbook(filename=excel_path, keep_links=False)
        self.__sheet = self.__book.worksheets[self.__sheet_index]
        self.attribute_map = self.reverse_column_map(self.column_map)
    
    def close(self):
        self.__book.save(self.__path)
        self.__book.close()

    def markup_with_errors(self):
        error_fill = PatternFill(fill_type='solid', start_color='FF0000')
        for row_index, row_errors in self.data.get_all_errors().items():
            error_count = 0
            for entity_type, entity_errors in row_errors.items():
                for attribute_name, attribute_errors in entity_errors.items():
                    error_count = error_count + len(attribute_errors)
                    human_errors = self.human_attribute_errors(entity_type, attribute_name, attribute_errors)
                    cell_index = self.lookup_cell_index(entity_type, attribute_name, row_index)
                    self.__sheet[cell_index].comment = self.__get_error_comment(human_errors, self.__sheet[cell_index].comment)
                    self.__sheet[cell_index].fill = error_fill
            if error_count:
                self.__sheet[f'A{row_index}'] = f'{error_count} Errors'
                self.__sheet[f'A{row_index}'].fill = error_fill

    # ToDo: When Submission accession functionality is done, save all 'new' accessions for all types
    def add_biosample_accessions(self):
        for entity in self.data.get_entities('sample'):
            if 'biosample' in entity.attributes:
                for row_index in self.data.get_rows_from_id(entity.identifier):
                    cell_index = self.lookup_cell_index('sample', 'sample_accession', row_index)
                    self.__sheet[cell_index] = entity.attributes['biosample']['accession']

    def lookup_cell_index(self, entity_type, attribute, row):
        attribute_key = f'{entity_type}.{attribute}'
        if attribute_key not in self.attribute_map:
            column_letter = 'C'
        else:
            column_letter = self.attribute_map[attribute_key]
        return f'{column_letter}{row}'

    @staticmethod
    def reverse_column_map(column_map: dict) -> dict:
        attribute_map = {}
        for letter, info in column_map.items():
            attribute_key = f"{info['object']}.{info['attribute']}"
            attribute_map[attribute_key] = letter
        return attribute_map

    @staticmethod
    def __clear_markup(excel_path, sheet_index):
        with closing(load_workbook(filename=excel_path, keep_links=False)) as book:
            sheet = book.worksheets[sheet_index]
            sheet.delete_cols(1, 1)
            sheet.insert_cols(1, 1)

            for row in sheet.iter_rows():
                for cell in row:
                    cell.fill = PatternFill()
                    cell.comment = None
            book.save(excel_path)

    @staticmethod
    def __get_error_comment(human_errors, existing_comment: Comment = None):
        stack = []
        if existing_comment and existing_comment.text:
            stack.append(existing_comment.text)
        stack.extend(human_errors)
        text = '\r\n'.join(stack)
        comment = Comment(text, f'Validation')
        comment.width = 500
        comment.height = 100
        return comment
