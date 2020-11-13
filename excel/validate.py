from typing import List

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.worksheet.worksheet import Worksheet
from validation.base import BaseValidator
from .load import ExcelLoader
from .clean import (object_has_attribute, clean_validation, clean_object, clean_key, clean_name,
                    clean_formula_list, clean_validation_list, valid_date)


def object_has_accession(object_data: dict):
    return (object_has_attribute(object_data, 'study_accession')
            or object_has_attribute(object_data, 'sample_accession'))


def validate_object(object_validation: dict, object_data: dict):
    errors = []
    for name, info in object_validation.items():
        if not object_has_attribute(object_data, name):
            if 'mandatory' in info and not object_has_accession(object_data):
                mandatory = info['mandatory'].strip()
                if mandatory == 'M':
                    errors.append(f'Error: Mandatory Attribute {name} is missing.')
                elif mandatory == 'R':
                    errors.append(f'Warning: Recommended Attribute {name} is missing.')
                elif mandatory != 'O':
                    errors.append(f'Info: Attribute {name} may be required. {mandatory}')
        else:
            value = object_data[name]
            if ('format' in info
                    and info['format'] == 'YYYY-MM-DD'
                    and not valid_date(value)):
                errors.append(f'Error: {name} has value {value}, which is not in date format YYYY-MM-DD.')
            if ('accepted_values' in info
                    and clean_validation(value) not in info['accepted_values']):
                accepted = info['accepted_values']
                errors.append(f'Error: {name} has value {value}, which is not in list of accepted values. {accepted}')
    if errors:
        object_data['errors'] = errors
    return errors


def validate_data_row(validation_map: dict, data_row):
    # ToDo: Refactor Best Guess validation to match output of schema_validation
    object_errors = []
    other_errors = []
    all_errors = []
    for object_name, validation_info in validation_map.items():
        if object_name not in data_row:
            object_errors.append(f'Error: Object {object_name} is missing.')
        else:
            other_errors.extend(validate_object(validation_info, data_row[object_name]))
    if object_errors:
        data_row.setdefault('errors', []).extend(object_errors)
    all_errors.extend(object_errors)
    all_errors.extend(other_errors)
    return all_errors


def validate_data_list(validation_map: dict, data):
    validation_report = {}
    for row_index, data_row in data.items():
        row_errors = validate_data_row(validation_map, data_row)
        if row_errors:
            validation_report[row_index] = row_errors
    return validation_report


def get_excel_validations(validations: DataValidationList):
    validation_list = {}
    # If Excel's data validation is used for displaying accepted values on header row,
    # Replace single selected value with all accepted values
    for validation in validations:
        if validation.validation_type == 'list':
            for cell_range in validation.ranges:
                if (cell_range.min_col == cell_range.max_col
                        and cell_range.min_row == 4
                        and cell_range.max_row == 4):
                    if ',' in validation.formula1:
                        # Accepted values encoded directly into validation formula
                        value = clean_formula_list(validation.formula1)
                    else:
                        value = [clean_key(validation.formula1)]
                    validation_list[cell_range.coord] = value
    return validation_list


def get_accepted_lists(worksheet: Worksheet) -> dict:
    defined_lists = {}
    for column in worksheet.columns:
        values = []
        for cell in column:
            if cell.value is not None:
                values.append(cell.value)
        if len(values) > 1:
            key = clean_key(values.pop(0))
            defined_lists[key] = clean_validation_list(values)
    return defined_lists


def merge_accepted_lists(excel_validations: dict, accepted_lists: dict):
    update = {}
    for coord, values in excel_validations.items():
        list_name = clean_key(values[0])
        if len(values) == 1 and list_name in accepted_lists:
            update[coord] = accepted_lists[list_name]
    excel_validations.update(update)
    return excel_validations


def get_validation_map(worksheet: Worksheet, excel_validations: dict = None) -> dict:
    validation = {}
    name = False
    for column in worksheet.iter_cols(min_col=2, max_row=5):
        cells = []
        for cell in column:
            cells.append(cell)

        object_cell = cells[0]
        attribute_cell = cells[1]
        mandatory_cell = cells[2]
        format_cell = cells[3]
        units_cell = cells[4]

        # Update Object Name otherwise use most recent Object found
        if object_cell.value is not None:
            name = clean_object(object_cell.value)
        if name:
            if attribute_cell.value is not None:
                attribute = clean_name(attribute_cell.value)
                column_object = {}

                if mandatory_cell.value is not None:
                    column_object['mandatory'] = mandatory_cell.value

                if validation and format_cell.coordinate in excel_validations:
                    column_object['accepted_values'] = excel_validations[format_cell.coordinate]
                elif format_cell.value is not None:
                    column_object['format'] = format_cell.value

                if units_cell.value is not None:
                    column_object['units'] = units_cell.value

                validation.setdefault(name, {})[attribute] = column_object
    return validation


def get_validation_dict_from_excel(file_path, sheet_index=0):
    workbook = load_workbook(filename=file_path, read_only=False, keep_links=False)
    worksheet = workbook.worksheets[sheet_index]
    try:
        validations = get_excel_validations(worksheet.data_validations.dataValidation)
        if 'Accepted_Values' in workbook.sheetnames:
            accepted_lists = get_accepted_lists(workbook['Accepted_Values'])
            merge_accepted_lists(validations, accepted_lists)
        return get_validation_map(worksheet, validations)
    finally:
        workbook.close()


def validate_dict_from_excel(file_path, data, sheet_index=0):
    validation_map = get_validation_dict_from_excel(file_path, sheet_index)
    return validate_data_list(validation_map, data)


class ValidatingExcel(ExcelLoader):
    def __init__(self, excel_path: str, sheet_index=0):
        super().__init__(excel_path, sheet_index)
        self.errors = {}

    def validate(self, validator: BaseValidator):
        self.errors = validator.validate_data(self.rows)

    @staticmethod
    def human_entity_errors(entity_type: str, errors: dict) -> List[str]:
        translated_messages = []
        for attribute_name, attribute_errors in errors.items():
            translated_messages.extend(
                ValidatingExcel.human_attribute_errors(entity_type, attribute_name, attribute_errors))
        return translated_messages

    @staticmethod
    def human_attribute_errors(entity_type: str, attribute_name: str, attribute_errors: List[str]) -> List[str]:
        translated = []
        for error in attribute_errors:
            if error.startswith('should have required property'):
                message = f'Error: {entity_type} {error}'.replace('"', '\'')
            else:
                message = f'Error: {entity_type}.{attribute_name} {error}'.replace('"', '\'')
            translated.append(message)
        return translated
