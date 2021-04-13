from typing import List

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.worksheet.worksheet import Worksheet
from excel_submission_broker.clean import (entity_has_attribute, clean_validation, clean_entity_name, clean_key,
                         clean_name, clean_formula_list, clean_validation_list, is_valid_date)
from submission_broker.submission.entity import Entity
from submission_broker.validation.base import BaseValidator


class ExcelValidator(BaseValidator):
    def __init__(self, file_path, sheet_index=0):
        self.validation_map = self.__get_validation(file_path, sheet_index)
        # ToDo: Accept param for number of header rows, columns / mapping of headers

    def validate_entity(self, entity: Entity):
        if entity.identifier.entity_type not in self.validation_map:
            return
        entity_validation = self.validation_map[entity.identifier.entity_type]
        for attribute_name, attribute_validation in entity_validation.items():
            attribute_errors = []
            if not entity_has_attribute(entity.attributes, attribute_name):
                if 'mandatory' in attribute_validation:
                    mandatory = attribute_validation['mandatory'].strip()
                    if mandatory == 'M':
                        attribute_errors.append(f'should have required property \'{attribute_name}\'.')
                    elif mandatory == 'R':
                        attribute_errors.append('recommended attribute is missing.')
                    elif mandatory != 'O':
                        attribute_errors.append(f'may be required: {mandatory}')
            else:
                attribute_errors.extend(ExcelValidator.validate_attribute(attribute_validation, entity.attributes[attribute_name]))
            if attribute_errors:
                entity.add_errors(attribute_name, attribute_errors)

    @staticmethod
    def validate_attribute(attribute_validation: dict, value) -> List[str]:
        attribute_errors = []
        if ('format' in attribute_validation
                and attribute_validation['format'] == 'YYYY-MM-DD'
                and not is_valid_date(value)):
            attribute_errors.append(f"'{value}' is not in date format: YYYY-MM-DD.")
        if ('accepted_values' in attribute_validation
                and clean_validation(value) not in attribute_validation['accepted_values']):
            accepted = attribute_validation['accepted_values']
            attribute_errors.append(f"'{value}' is not in list of accepted values: {accepted}")
        return attribute_errors

    @staticmethod
    def __get_validation(file_path, sheet_index=0):
        workbook = load_workbook(filename=file_path, read_only=False, keep_links=False)
        worksheet = workbook.worksheets[sheet_index]
        try:
            validations = ExcelValidator.__get_excel_validations(worksheet.data_validations.dataValidation)
            if 'Accepted_Values' in workbook.sheetnames:
                accepted_lists = ExcelValidator.__get_accepted_lists(workbook['Accepted_Values'])
                ExcelValidator.__merge_accepted_lists(validations, accepted_lists)
            return ExcelValidator.__load_validation_map(worksheet, validations)
        finally:
            workbook.close()
    
    @staticmethod
    def __get_excel_validations(validations: DataValidationList):
        validation_list = {}
        # If Excel's data validation is used for displaying accepted values on header row,
        # Replace single selected value with all accepted values
        for validation in validations:
            if validation.validation_type == 'list':
                for cell_range in validation.ranges:
                    if (cell_range.min_col == cell_range.max_col
                            and cell_range.min_row == 4
                            and cell_range.max_row == 4):
                        if ',' in validation.formula1:
                            # Accepted values encoded directly into validation formula
                            value = clean_formula_list(validation.formula1)
                        else:
                            value = [clean_key(validation.formula1)]
                        validation_list[cell_range.coord] = value
        return validation_list
    
    @staticmethod
    def __get_accepted_lists(worksheet: Worksheet) -> dict:
        defined_lists = {}
        for column in worksheet.columns:
            values = []
            for cell in column:
                if cell.value is not None:
                    values.append(cell.value)
            if len(values) > 1:
                key = clean_key(values.pop(0))
                defined_lists[key] = clean_validation_list(values)
        return defined_lists

    @staticmethod
    def __merge_accepted_lists(excel_validations: dict, accepted_lists: dict):
        update = {}
        for coord, values in excel_validations.items():
            list_name = clean_key(values[0])
            if len(values) == 1 and list_name in accepted_lists:
                update[coord] = accepted_lists[list_name]
        excel_validations.update(update)
        return excel_validations
    
    @staticmethod
    def __load_validation_map(worksheet: Worksheet, excel_validations: dict = None) -> dict:
        validation = {}
        name = False
        for column in worksheet.iter_cols(min_col=2, max_row=5):
            cells = []
            for cell in column:
                cells.append(cell)

            object_cell = cells[0]
            attribute_cell = cells[1]
            mandatory_cell = cells[2]
            format_cell = cells[3]
            units_cell = cells[4]

            # Update Object Name otherwise use most recent Object found
            if object_cell.value is not None:
                name = clean_entity_name(object_cell.value)
            if name:
                if attribute_cell.value is not None:
                    attribute = clean_name(attribute_cell.value)
                    column_object = {}

                    if mandatory_cell.value is not None:
                        column_object['mandatory'] = mandatory_cell.value

                    if validation and format_cell.coordinate in excel_validations:
                        column_object['accepted_values'] = excel_validations[format_cell.coordinate]
                    elif format_cell.value is not None:
                        column_object['format'] = format_cell.value

                    if units_cell.value is not None:
                        column_object['units'] = units_cell.value

                    validation.setdefault(name, {})[attribute] = column_object
        return validation
