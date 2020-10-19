from openpyxl import load_workbook


def clean_object(name: str) -> str:
    return clean_name(name.partition('-')[0])


def clean_name(name: str) -> str:
    return name.strip().lower().replace(' ', '_').replace('/', '_')


def get_column_map(worksheet) -> dict:
    # Uses iter_rows for faster reads, requires workbook read_only=True
    column_map = {}
    header_rows = []
    object_name = False
    for row in worksheet.iter_rows(min_col=2, max_row=5):
        header_rows.append(row)
    for column_index in range(0, len(header_rows[0])):
        object_cell = header_rows[0][column_index]
        attribute_cell = header_rows[1][column_index]
        mandatory_cell = header_rows[2][column_index]
        format_cell = header_rows[3][column_index]
        units_cell = header_rows[4][column_index]
        column_info = {}

        # Update Object Name otherwise use most recent Object found
        if object_cell.value is not None:
            object_name = clean_object(object_cell.value)
        if object_name:
            column_info['object'] = object_name
        if mandatory_cell.value is not None:
            column_info['mandatory'] = mandatory_cell.value
        if format_cell.value is not None:
            column_info['format'] = format_cell.value
        if units_cell.value is not None:
            column_info['units'] = units_cell.value
        if attribute_cell.value is not None:
            column_info['attribute'] = clean_name(attribute_cell.value)
            column_map[attribute_cell.column_letter] = column_info
    return column_map


def get_data(worksheet, column_map: dict = None):
    if column_map is None:
        column_map = get_column_map(worksheet)
    data = []
    # Import cell values into data object
    # Uses .iter_rows for faster reads, requires workbook read_only=True
    for row in worksheet.iter_rows(min_row=6, min_col=2):
        row_data = {}
        for cell in row:
            if cell.value is not None:
                column_info = column_map[cell.column_letter]
                if column_info['object'] not in row_data:
                    row_data[column_info['object']] = {}
                row_data[column_info['object']][column_info['attribute']] = str(cell.value)
        if len(row_data) > 0:
            data.append(row_data)
    return data


def get_dict_from_excel(file_path, sheet_index=0):
    workbook = load_workbook(filename=file_path, read_only=True, keep_links=False)
    try:
        worksheet = workbook.worksheets[sheet_index]
        return get_data(worksheet)
    finally:
        workbook.close()
